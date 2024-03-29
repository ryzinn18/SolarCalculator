# SolarCalculator/src/backend/inputs.py
# Handles an input type by calling the correct input function and validates InputData
from pydantic import PositiveFloat

from logging import getLogger
from typing import Callable
from datetime import datetime as dt

from utils import IntListMonthly, InputData, Status, MONTHS_MAP

LOGGER = getLogger(__name__)


class InputError(Exception):
    """Custom exception for an invalid Input keyword"""
    valid_input_types = ['csv', 'xlsx', 'sheet', 'form']


def validate(function: Callable) -> Callable:
    """Decorator function for handling exceptions and logging for the input functions."""

    def wrapper_validate(*args, **kwargs):
        """Wraps function call in a try/except clause to catch FileNotFound, ValueError and general Exceptions"""
        LOGGER.info('Validating input data.')
        try:
            result = function(*args, **kwargs)
        except FileNotFoundError as e:
            LOGGER.error(f'The file specified does not exist.', exc_info=True)
            raise FileNotFoundError(
                f"The file specified at the below location does not exist:\n"
                + f"\t{e.filename}\n"
            )
        except ValueError:
            LOGGER.error(f'The input data could not be validated.', exc_info=True)
            raise ValueError(
                f"One or more of the values entered are invalid.\n"
                + "Please review the values entered and try again."
            )
        except TypeError:
            LOGGER.error(f'One or more of the data types passed was invalid.', exc_info=True)
            raise TypeError(
                "One or more of the values entered is of an invalid data type.\n"
                + "Please review the values entered and try again."
            )
        except Exception as e:
            LOGGER.error(e, exc_info=True)
            raise e
        else:
            LOGGER.info('Input data successfully validated.')
            return result

    return wrapper_validate


def _validate_mod_kwh(in_data: str) -> float:
    """
    Helper function to validate that the data passed can be converted to a float and is between 0 and 1.5.

    :param in_data: A string that should be able to be converted to a float.
    :return the float value if a valid input given.
    :raise ValueError if in_data cannot be converted to a float or the float value is not between 0 and 1.5.
    """

    LOGGER.info(f'Validating {in_data} for the mod kWh value')

    # Validate the value passed is numerical.
    try:
        result = float(in_data)
    except ValueError:
        LOGGER.exception(f'The value {in_data} is invalid (not a decimal between 0 and 1) for mod kWh', exc_info=True)
        raise ValueError(
            "The value entered for solar module capacity must be numerical."
        )
    # Validate the value is greater than 0 and less than 1.5 and is invalid
    if 0 < result <= 1.5:
        LOGGER.info(f'The value {result} is valid for mod kWh')
        return result
    else:
        LOGGER.exception(f'The value {result} is not between 0 and 1.5 and is invalid')
        raise ValueError(
            "The value provided is is not greater than 0 and less than 1.5 and is invalid."
        )


def _calculate_cost_per_kwh(cost: IntListMonthly, consumption: IntListMonthly) -> PositiveFloat:
    """Return the average cost per kWh for a year's inputs"""

    # Get a list of the cost/kWh for each month
    monthly_cost_per_kwh = [(cos / kwh) for cos, kwh in zip(cost, consumption)]
    # Calculate and return the average cost/kWh for the year
    return round(sum(monthly_cost_per_kwh) / len(monthly_cost_per_kwh), 2)


def _clean_name(name: str) -> str:
    """Remove all spaces and non-alphanumeric characters from name."""
    return ''.join(s for s in name if s.isalnum())


@validate
def input_csv(file_path: str, time_stamp: str) -> InputData:
    """Read a csv at specified path and return InputData object."""

    from csv import reader

    LOGGER.info(f'Collecting input data from csv at following location: {file_path}')

    def _read_data(r_min: int, r_max: int, column: int, convert: bool, round_to: int = None) -> list:
        """Read data from input csv based on range min and max values. Converts str to float if specified."""

        # Create output list
        output = []
        # Open the file in the context manager
        with open(file_path, 'r') as file:
            csv = reader(file)
            # Get input consumption and cost data from csv
            data_rows = [row for i, row in enumerate(csv) if i in range(r_min, r_max)]
            # Write the data in the column specified to the output list
            for row in data_rows:
                output.append(row[column])

        # Convert values to floats and round them if convert is True, else return the output as is
        return [round(float(val), round_to) for val in output] if convert else output

    # Get data from input csv file
    consumption = _read_data(r_min=1, r_max=13, column=2, convert=True, round_to=0)
    cost = _read_data(r_min=1, r_max=13, column=3, convert=True, round_to=2)
    user_data = _read_data(r_min=14, r_max=17, column=2, convert=False)

    # Instantiate InputData model
    result = InputData(
        uid=_clean_name(user_data[0]) + time_stamp,
        name=_clean_name(name=user_data[0]),
        time_stamp=dt.now().__str__(),
        status=Status(status_code=200, message="get_inputs() called input_csv() successfully."),
        address=user_data[1],
        mod_kwh=_validate_mod_kwh(in_data=user_data[2]),
        consumption_monthly=consumption,
        consumption_annual=sum(consumption),
        cost_monthly=cost,
        cost_annual=round(sum(cost), 2),
        cost_per_kwh=_calculate_cost_per_kwh(cost=cost, consumption=consumption)
    )

    LOGGER.info(f'InputData successfully collected from csv: {file_path}')
    return result


@validate
def input_xlsx(file_path: str, time_stamp: str) -> InputData:
    """Read a xlsx at specified path and return InputData object."""

    from openpyxl import load_workbook

    LOGGER.info(f'Collecting input data from xlsx at following location: {file_path}')

    def _read_data(r_min: int, r_max: int, column: str, convert: bool, round_to: int = None) -> list:
        """Read data from input xlsx based on range min and max values. Converts str to float if specified."""

        # Create output list
        output = []
        for i in range(r_min, r_max):
            output.append(sheet[f'{column}{i}'].value)

        # Convert values to floats and round them if convert is True, else return the output as is
        return [round(float(val), round_to) for val in output] if convert else output

    # Create sheet object from load_workbook
    sheet = load_workbook(file_path)['Sheet1']

    # Get data from input xlsx file
    consumption = _read_data(r_min=2, r_max=14, column='C', convert=True, round_to=0)
    cost = _read_data(r_min=2, r_max=14, column='D', convert=True, round_to=2)
    user_data = _read_data(r_min=15, r_max=18, column='C', convert=False)

    # Instantiate InputData model
    result = InputData(
        uid=_clean_name(user_data[0]) + time_stamp,
        name=_clean_name(name=user_data[0]),
        time_stamp=dt.now().__str__(),
        status=Status(status_code=200, message="get_inputs() called input_xlsx() successfully."),
        address=user_data[1],
        mod_kwh=_validate_mod_kwh(in_data=user_data[2]),
        consumption_monthly=consumption,
        consumption_annual=sum(consumption),
        cost_monthly=cost,
        cost_annual=round(sum(cost), 2),
        cost_per_kwh=_calculate_cost_per_kwh(cost=cost, consumption=consumption)
    )

    LOGGER.info(f'InputData successfully collected from xlsx: {file_path}')
    return result


# @validate
def input_sheets(sheet_id: str, time_stamp: str) -> InputData:
    """
    _authenticate_google_api_token() script from below and updated slightly by me.
    https://medium.com/analytics-vidhya/how-to-read-and-write-data-to-google-spreadsheet-using-python-ebf54d51a72c

    Will redirect to a google authenticate page if you have not run in a while and ask you to manually authenticate.
        Should find a way around this. Probably a permissions issue.
    """

    from googleapiclient.discovery import build, Resource
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from google.auth.exceptions import RefreshError
    from os import remove as os_remove
    from os.path import exists as os_exists
    import pickle

    LOGGER.info(f'Collecting input data from the Google Sheet with the following sheet id: {sheet_id}')

    scopes = ['https://www.googleapis.com/auth/spreadsheets']

    def _authenticate_google_api_token() -> Resource:
        """Helper function to authenticate a google sheets api token"""

        LOGGER.info('Attempting to authenticate the google sheets api')

        # 1. set credentials to None
        credentials = None
        # 2. if token.pickle exists, open it and set credentials to it
        if os_exists('./.creds/token.pickle'):
            LOGGER.info('token.pickle object found')
            with open('./.creds/token.pickle', 'rb') as token:
                credentials = pickle.load(token)
        # 3. if token.pickle did not exist or was invalid do 3.a or 3.b then recreate the token.pickle
        if not credentials or not credentials.valid:
            LOGGER.info('Credentials token.pickle found but invalid, attempting to resolve')
            # 3.a if you have a credentials object but they're expired and support refreshing \
            #     try to refresh and if that fails then log this error
            if credentials and credentials.expired and credentials.refresh_token:
                try:
                    credentials.refresh(Request())
                    LOGGER.info('Refreshing the token.pickle successful')
                except RefreshError:
                    LOGGER.error("Credentials could not be refreshed, deleting token.pickle and trying again")
                    os_remove('./.creds/token.pickle')
                    # Recursive call to rerun the process once the token is deleted. Max call stack = 1
                    _authenticate_google_api_token()
            # 3.b in all other cases, create a new flow and token.pickle
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    r'./.creds/credentials.json', scopes)
                credentials = flow.run_local_server(port=0)
            # Finally, recreate the token.pickle
            with open('./.creds/token.pickle', 'wb') as token:
                pickle.dump(credentials, token)

        # 4. Build and return a service (Resource) object
        LOGGER.info('Successfully authenticated the google sheets api')
        return build('sheets', 'v4', credentials=credentials)

    def _read_data(column: int, convert: bool, round_to: int = None) -> list:
        """Read data from input google sheet. Converts str to float if specified."""

        # Get data from specified column
        output = input_data.get('valueRanges')[column].get('values')
        # Convert values to floats and round them if convert is True, else return the output as is
        return [round(float(sublist[0]), round_to) for sublist in output] if convert else output

    # Call the Sheets API and get raw data
    sheet_ranges = ['C2:C13', 'D2:D13', 'C15:C17']
    service = _authenticate_google_api_token()
    sheet = service.spreadsheets()
    input_data = sheet.values().batchGet(spreadsheetId=sheet_id, ranges=sheet_ranges).execute()

    # Get data from input google sheet
    user_data = _read_data(column=2, convert=False)
    consumption = _read_data(column=0, convert=True, round_to=0)
    cost = _read_data(column=1, convert=True, round_to=2)

    # Instantiate InputData model
    result = InputData(
        uid=_clean_name(user_data[0][0]) + time_stamp,
        name=_clean_name(user_data[0][0]),
        time_stamp=dt.now().__str__(),
        status=Status(status_code=200, message="get_inputs() called input_sheets() successfully."),
        address=user_data[1][0],
        mod_kwh=_validate_mod_kwh(in_data=user_data[2][0]),
        consumption_monthly=consumption,
        consumption_annual=sum(consumption),
        cost_monthly=cost,
        cost_annual=round(sum(cost), 2),
        cost_per_kwh=_calculate_cost_per_kwh(cost=cost, consumption=consumption)
    )

    LOGGER.info(f'InputData successfully collected from Google Sheet')
    return result


@validate
def input_form(input_obj: dict) -> InputData:
    """Takes web input object and returns InputData object."""

    LOGGER.info(f'Collecting input data from web input object: {input_obj}')
    consumption_monthly = [input_obj.get("monthly_data")[month][0] for month in MONTHS_MAP.values()]
    cost_monthly = [input_obj.get("monthly_data")[month][1] for month in MONTHS_MAP.values()]

    result = InputData(
        uid=input_obj.get("uid"),
        name=input_obj.get("name"),
        time_stamp=dt.now().__str__(),
        status=Status(status_code=200, message="get_inputs() called input_form() successfully."),
        address=input_obj.get('address'),
        mod_kwh=input_obj.get('mod_kwh'),
        consumption_monthly=consumption_monthly,
        consumption_annual=round(sum(consumption_monthly)),
        cost_monthly=cost_monthly,
        cost_annual=round(sum(cost_monthly), 2),
        cost_per_kwh=_calculate_cost_per_kwh(
            cost=cost_monthly,
            consumption=consumption_monthly
        )
    )
    LOGGER.info(f'InputData successfully collected from web input object: {input_obj}')
    return result


def get_inputs(input_event: dict) -> InputData:
    """Calls correct input function or raises an InputError. Will catch, log and raise Exceptions."""

    LOGGER.info(f'The input_handler() has been called for event: {input_event}')

    input_type = input_event.get('type')
    LOGGER.info(f'The event type has been found to be: {input_type}')

    # If/elif/else clause to call correct handler function. Raise and log error if no input function is called.
    if input_type == 'csv':
        input_data = input_csv(file_path=input_event['csv_source'])
    elif input_type == 'xlsx':
        input_data = input_xlsx(file_path=input_event['xlsx_source'])
    elif input_type == 'sheet':
        input_data = input_sheets(sheet_id=input_event['sheet_id'])
    elif input_type == 'form':
        input_data = input_form(input_obj=input_event['form'])
    else:
        LOGGER.error(f'The input type passed is invalid: {input_type}')
        raise InputError(
            f"The input '{input_type}' is invalid.\n"
            + f"You must enter one of the following input types: {*InputError.valid_input_types,}"
        )

    LOGGER.info(f'InputData successfully received via {input_type} for {input_data.name}')
    LOGGER.info(input_data)
    return input_data


def input_handler(event: dict) -> dict:
    """Handler Function to call get_inputs() and return validated object."""

    LOGGER.info(f'Called Handler function for getting input data event from event: {event}')

    # Handle getting the necessary data
    try:
        input_data = get_inputs(
            input_event=event,
        ).dict()
        LOGGER.info(f'Lambda Handler for input data successfully executed for uid: {input_data.get("uid")}')
    except Exception as e:
        time_stamp = dt.now().__str__()
        uid = event.get("uid") if event.get("uid") else "NA" + time_stamp
        input_data = {
            "uid": uid,
            "time_stamp": time_stamp,
            "status": Status(
                status_code=400,
                message=f"get_inputs() called unsuccessfully due to error: {e.__repr__()}"
            )
        }
        LOGGER.error(e, exc_info=True)

    return input_data


if __name__ == '__main__':
    from utils import import_json, SAMPLES
    print(input_handler(import_json(SAMPLES['event_valid_form'])).get("cost_monthly"))
